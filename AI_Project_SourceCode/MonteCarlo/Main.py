import numpy as np
import utilityMethods as um
import xlrd
import math
from math import ceil
from math import floor
import openpyxl
import random

class TrafficStateTuple:
    '''Initializing the state of a traffic Light'''
    laneLength=10
    def __init__(self,queue1,queue2,phaseTime1,phaseTime2):
        self.queue1=queue1
        self.queue2=queue2
        self.phaseTime1=phaseTime1
        self.phaseTime2=phaseTime2

actionSpace=[]
for i in range(1,13):
    actionSpace.append(i/12)

gamma=0.8
epsilon=0.3
flag = 0


def selectAction(state):
    location = ("QMatrix.xlsx")
    openWb = xlrd.open_workbook(location)
    sheet = openWb.sheet_by_index(0)
    unexplored = []
    decayRate=np.random.uniform(0,1)
    if decayRate < epsilon:
        print("Action random for intersection 1")
        for i in range(12):
            if sheet.cell_value(state,i)==0:
                unexplored.append(i+1)
        if len(unexplored)!=0:
            print("length:",len(unexplored))
            select = random.randint(0,len(unexplored))

            action = unexplored[select]
        else:
            action = np.random.randint(1, 13)
    else:
        print("Action optimal for intersection 1")
        action = um.findMaxQvalueAction(state, "QMatrix.xlsx")
    return action

def selectActionFor2(state):
    location = ("QMatrix.xlsx")
    openWb = xlrd.open_workbook(location)
    sheet = openWb.sheet_by_index(0)
    unexplored = []
    decayRate=np.random.uniform(0,1)
    if decayRate < epsilon:
        print("Action random for intersection 2")
        for i in range(12):
            if sheet.cell_value(state,i)==0:
                unexplored.append(i+1)
        if len(unexplored)!=0:
            print("length:", len(unexplored))
            select = random.randint(0,len(unexplored)-1)
            action = unexplored[select]
        else:
            action = np.random.randint(1, 13)
    else:
        print("Action optimal for intersection 2")
        action = um.findMaxQvalueAction(state, "QMatrix2.xlsx")
    return action



def forIntersection1():
    global flag
    '''get output matrix from simul'''
    location=("traffic.xls")
    openWb = xlrd.open_workbook(location)
    sheet = openWb.sheet_by_index(0)

    locationB=("traffic2.xls")
    openWbB = xlrd.open_workbook(locationB)
    sheetB = openWbB.sheet_by_index(0)

    '''get Prev State xlsx for intersection 1'''
    location2=("PreviousState.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)

    '''get Prev State xlsx for intersection 2'''
    location2B=("PreviousState.xlsx")
    openWb2B = xlrd.open_workbook(location2B)
    sheet2B = openWb2B.sheet_by_index(0)

    file=openpyxl.load_workbook('PreviousState.xlsx')
    sheet3=file['Sheet1']

    file2=openpyxl.load_workbook('PreviousState2.xlsx')
    sheetT2=file2['Sheet1']
    #print(sheet.cell_value(0,0),sheet.cell_value(0,1),sheet.cell_value(0,2),sheet.cell_value(0,3))
    sum1=0
    sum2=0
    sum3=0
    sum4=0
    #print((sheet.cell_value(0,0)))
    for i in range(min(sheet.nrows,sheetB.nrows)):
        sum1=sum1+int(sheet.cell_value(i,1))
        sum2=sum2+int(sheet.cell_value(i,3))
        sum3=sum3+int(sheet.cell_value(i,5))
        sum4=sum4+int(sheetB.cell_value(i,7))
    avgQueue1=sum1/sheet.nrows
    avgQueue2=sum2/sheet.nrows
    avgUpstream=sum3/sheet.nrows
    avgDownstream=sum4/sheetB.nrows

    if um.findRow("PreviousState.xlsx")==1:

        currentTrafficState = TrafficStateTuple(math.ceil(avgQueue1), math.ceil(avgQueue2),sheet.cell_value(0, 0),sheet.cell_value(0, 2))
        prevTrafficState = TrafficStateTuple(sheet.cell_value(1, 1), sheet.cell_value(1, 3),sheet.cell_value(0, 0),sheet.cell_value(0, 2))
        state=um.trafficState(currentTrafficState,prevTrafficState)
        action=selectAction(state)
        oldTime1=sheet.cell_value(0, 0)
        oldTime2=sheet.cell_value(0, 2)
        #print(oldTime1,oldTime2)
        '''update PreviousState.xlsx with state,queue1,queue2,newP1,newP2,action,avgUpstream'''
        updateRow=str(um.findRow("PreviousState.xlsx"))
        updateRowForT2=str(um.findRow("PreviousState2.xlsx"))
        A='A'+updateRow
        B='B'+updateRow
        C='C'+updateRow
        D='D'+updateRow
        E='E'+updateRow
        F='F'+updateRow
        G='G'+updateRow

        #print(A,B,C,D,E,F)
        sheet3[A]=state
        sheet3[B]=math.ceil(avgQueue1)
        sheet3[C]=math.ceil(avgQueue2)
        sheet3[F]=action
        sheet3[G]=avgUpstream

        file.save('PreviousState.xlsx')
        x,y=um.calculatePhaseForA(action,math.ceil(avgQueue1),math.ceil(avgQueue2),oldTime1,oldTime2,avgDownstream)
        sheet3[D]=ceil(x)
        sheet3[E]=ceil(y)
        file.save('PreviousState.xlsx')

        print("For intersection 1:",ceil(x),ceil(y))
    else:
        currentTrafficState = TrafficStateTuple(math.ceil(avgQueue1), math.ceil(avgQueue2),sheet.cell_value(0, 0),sheet.cell_value(0, 2))
        prevTrafficState = TrafficStateTuple(sheet2.cell_value(sheet2.nrows-1, 1), sheet2.cell_value(sheet2.nrows-1, 2),sheet2.cell_value(sheet2.nrows-1,3),sheet2.cell_value(sheet2.nrows-1,4))
        #print(prevTrafficState.queue1,prevTrafficState.queue2,prevTrafficState.phaseTime1,prevTrafficState.phaseTime2)
        reward = um.calculateReward(currentTrafficState, prevTrafficState)
        state=um.trafficState(currentTrafficState,prevTrafficState)
        #um.updateA(state,currentTrafficState,prevTrafficState)

        action=selectAction(state)
        oldTime1=sheet2.cell_value(sheet2.nrows-1,3)
        oldTime2=sheet2.cell_value(sheet2.nrows-1,4)
        updateRow = str(um.findRow('PreviousState.xlsx'))
        updateRow2 = str(um.findRow("PreviousState.xlsx") - 1)

        A='A'+updateRow
        B='B'+updateRow
        C='C'+updateRow
        D='D'+updateRow
        E='E'+updateRow
        F='F'+updateRow
        G = 'G'+updateRow
        H = 'H'+updateRow2

        #print(A,B,C,D,E,F)
        sheet3[A]=state
        sheet3[B]=math.ceil(avgQueue1)
        sheet3[C]=math.ceil(avgQueue2)

        sheet3[F]=action
        sheet3[G] = avgUpstream
        file.save('PreviousState.xlsx')
        x,y=um.calculatePhaseForA(action,math.ceil(avgQueue1),math.ceil(avgQueue2),oldTime1,oldTime2,avgDownstream)
        sheet3[D]=ceil(x)
        sheet3[E]=ceil(y)
        sheet3[H] = reward
        file.save('PreviousState.xlsx')
        '''update PreviousState.xlsx with state,queue1,queue2,newP1,newP2,action,avgUpstream'''
        print("For intersection 1:",ceil(x),ceil(y))
        '''check for convergence'''
        if avgQueue1+avgQueue2/28 <= 0.3 or um.findRow('PreviousState.xlsx') == 10:
            flag = 1


def forIntersection2():
    global flag
    '''get output matrix from simul'''
    location = ("traffic2.xls")
    openWb = xlrd.open_workbook(location)
    sheet = openWb.sheet_by_index(0)

    locationA=("traffic.xls")
    openWbA = xlrd.open_workbook(locationA)
    sheetA = openWbA.sheet_by_index(0)

    '''get Prev State xlsx'''
    location2 = ("PreviousState2.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)

    '''get Prev State xlsx for intersection 1'''
    location2A=("PreviousState.xlsx")
    openWb2A = xlrd.open_workbook(location2A)
    sheet2A = openWb2A.sheet_by_index(0)

    file = openpyxl.load_workbook('PreviousState2.xlsx')
    sheet3 = file['Sheet1']

    # print(sheet.cell_value(0,0),sheet.cell_value(0,1),sheet.cell_value(0,2),sheet.cell_value(0,3))
    sum1 = 0
    sum2 = 0
    sum3=0
    sum4=0
    # print((sheet.cell_value(0,0)))
    for i in range(min(sheet.nrows,sheetA.nrows)):
        sum1 = sum1 + int(sheet.cell_value(i, 1))
        sum2 = sum2 + int(sheet.cell_value(i, 3))
        sum3=sum3+int(sheet.cell_value(i,5))
        sum4=sum4+int(sheetA.cell_value(i,7))
    avgQueue1 = sum1 / sheet.nrows
    avgQueue2 = sum2 / sheet.nrows
    avgDownstream=sum3/sheet.nrows
    avgUpstream=sum4/sheetA.nrows

    if um.findRow('PreviousState2.xlsx') == 1:

        currentTrafficState = TrafficStateTuple(math.ceil(avgQueue1), math.ceil(avgQueue2), sheet.cell_value(0, 0),
                                                sheet.cell_value(0, 2))
        prevTrafficState = TrafficStateTuple(sheet.cell_value(1, 1), sheet.cell_value(1, 3), sheet.cell_value(0, 0),
                                             sheet.cell_value(0, 2))
        state = um.trafficState(currentTrafficState, prevTrafficState)
        action = selectActionFor2(state)
        oldTime1 = sheet.cell_value(0, 0)
        oldTime2 = sheet.cell_value(0, 2)
        #print(oldTime1,oldTime2)
        #print("int2")
        '''update PreviousState.xlsx with state,queue1,queue2,newP1,newP2,action,avgDownstream'''
        updateRow = str(um.findRow('PreviousState2.xlsx'))
        A = 'A' + updateRow
        B = 'B' + updateRow
        C = 'C' + updateRow
        D = 'D' + updateRow
        E = 'E' + updateRow
        F = 'F' + updateRow
        G = 'G' + updateRow
        # print(A,B,C,D,E,F)
        sheet3[A] = state
        sheet3[B] = math.ceil(avgQueue1)
        sheet3[C] = math.ceil(avgQueue2)

        sheet3[F] = action
        sheet3[G]=avgDownstream
        file.save('PreviousState2.xlsx')
        x, y = um.calculatePhaseForB(action, math.ceil(avgQueue1), math.ceil(avgQueue2), oldTime1, oldTime2,avgUpstream)
        sheet3[D] = ceil(x)
        sheet3[E] = ceil(y)
        file.save('PreviousState2.xlsx')
        print("For intersection 2:",ceil(x), ceil(y))
    else:
        #print("else of 2")
        currentTrafficState = TrafficStateTuple(math.ceil(avgQueue1), math.ceil(avgQueue2), sheet.cell_value(0, 0),
                                                sheet.cell_value(0, 2))
        prevTrafficState = TrafficStateTuple(sheet2.cell_value(sheet2.nrows - 1, 1),
                                             sheet2.cell_value(sheet2.nrows - 1, 2),
                                             sheet2.cell_value(sheet2.nrows - 1, 3),
                                             sheet2.cell_value(sheet2.nrows - 1, 4))
        # print(prevTrafficState.queue1,prevTrafficState.queue2,prevTrafficState.phaseTime1,prevTrafficState.phaseTime2)

        state = um.trafficState(currentTrafficState, prevTrafficState)
        #um.updateB(state, currentTrafficState, prevTrafficState)
        reward = um.calculateReward(currentTrafficState, prevTrafficState)
        action = selectActionFor2(state)
        oldTime1 = sheet2.cell_value(sheet2.nrows - 1, 3)
        oldTime2 = sheet2.cell_value(sheet2.nrows - 1, 4)
        #print(oldTime1,oldTime2)
        updateRow = str(um.findRow('PreviousState2.xlsx'))
        updateRow2 = str(um.findRow("PreviousState2.xlsx") - 1)
        A = 'A' + updateRow
        B = 'B' + updateRow
        C = 'C' + updateRow
        D = 'D' + updateRow
        E = 'E' + updateRow
        F = 'F' + updateRow
        G = 'G' + updateRow
        H = 'H' + updateRow2
        # print(A,B,C,D,E,F)
        sheet3[A] = state
        sheet3[B] = math.ceil(avgQueue1)
        sheet3[C] = math.ceil(avgQueue2)

        sheet3[F] = action
        sheet3[G]=avgDownstream
        file.save('PreviousState2.xlsx')

        x, y = um.calculatePhaseForB(action, math.ceil(avgQueue1), math.ceil(avgQueue2), oldTime1, oldTime2,avgUpstream)
        sheet3[D] = ceil(x)
        sheet3[E] = ceil(y)
        sheet3[H] = reward
        file.save('PreviousState2.xlsx')
        '''update PreviousState.xlsx with state,queue1,queue2,newP1,newP2,action,avgDownstream2'''
        print("For intersection 2:",ceil(x), ceil(y))
        print(um.findRow('PreviousState2.xlsx'))
        if ((avgQueue1+avgQueue2/28) <= 0.3) or um.findRow('PreviousState2.xlsx') == 10:
            print("if")
            flag = 1



forIntersection1()
forIntersection2()
'''update Q Matrix'''
if flag == 1:
    print("Saturation reached.Updation begins.")
    location = ('PreviousState.xlsx')
    openWb = xlrd.open_workbook(location)
    sheet3 = openWb.sheet_by_index(0)
    file = openpyxl.load_workbook('QMatrix.xlsx')
    sheet = file['Sheet1']
    location2 = ("QMatrix.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)
    G = 0
    t = um.findRow('PreviousState.xlsx')-3
    state_action_pair = dict()
    while t>=0:
        s = int(sheet3.cell_value(t, 0))
        a = int(sheet3.cell_value(t, 5))
        search = str(s) + str(a)
        if search in state_action_pair.keys():
            val = state_action_pair[search]
            val[2] = val[2] + 1
        else:
            var = []
            var.append(s)
            var.append(a)
            var.append(1)
            state_action_pair[search] = var
        t =t -1
    #print(state_action_pair)
    t = um.findRow('PreviousState.xlsx')-3
    while t >= 0:
        G = (G * gamma) + sheet3.cell_value(t, 7)
        s = sheet3.cell_value(t, 0)
        a = sheet3.cell_value(t, 5)
        c = chr(int(a + 64)) + str(int(s))
        c1 = chr(int(a + 64)) + str(int(s + 9))
        st = str(int(s)) + str(int(a))
        print(c,c1)
        print((1/(int(sheet[c1].value)+state_action_pair[st][2])*(G-sheet[c].value)))
        sheet[c] = sheet[c].value + (1/(int(sheet[c1].value)+state_action_pair[st][2])*(G-sheet[c].value))
        sheet[c1] = int(sheet[c1].value)+state_action_pair[st][2]
        file.save('QMatrix.xlsx')
        t = t - 1
    '''for intersection 2'''
    location = ('PreviousState2.xlsx')
    openWb = xlrd.open_workbook(location)
    sheet3 = openWb.sheet_by_index(0)
    file = openpyxl.load_workbook('QMatrix2.xlsx')
    sheet = file['Sheet1']
    location2 = ("QMatrix2.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)
    G = 0
    t = um.findRow('PreviousState2.xlsx') - 3
    state_action_pair = dict()
    while t >= 0:
        s = int(sheet3.cell_value(t, 0))
        a = int(sheet3.cell_value(t, 5))
        search = str(s) + str(a)
        if search in state_action_pair.keys():
            val = state_action_pair[search]
            val[2] = val[2] + 1
        else:
            var = []
            var.append(s)
            var.append(a)
            var.append(1)
            state_action_pair[search] = var
        t = t - 1
    t = um.findRow('PreviousState2.xlsx') - 3
    print(t)
    while t >= 0:
        G = (G * gamma) + sheet3.cell_value(t, 7)
        s = sheet3.cell_value(t, 0)
        a = sheet3.cell_value(t, 5)

        c = chr(int(a + 64)) + str(int(s))
        c1 = chr(int(a + 64)) + str(int(s + 9))
        st = str(int(s)) + str(int(a))
        print(c, c1)
        print((1 / (int(sheet[c1].value) + state_action_pair[st][2]) * (G - sheet[c].value)))
        sheet[c] = sheet[c].value + (1 / (int(sheet[c1].value) + state_action_pair[st][2]) * (G - sheet[c].value))
        sheet[c1] = int(sheet[c1].value) + state_action_pair[st][2]
        file.save('QMatrix2.xlsx')
        t = t - 1

#print(avgQueue1,avgQueue2)



