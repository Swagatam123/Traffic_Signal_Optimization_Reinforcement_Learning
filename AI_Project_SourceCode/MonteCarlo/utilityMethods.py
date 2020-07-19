import numpy as np
from math import ceil
import TrafficLight
import openpyxl
from math import floor
import xlrd
gamma=0.8
alpha=0.1
def trafficState(currentTrafficState,prevTrafficState):
    queue=(currentTrafficState.queue1+currentTrafficState.queue2)/20
    flux=((currentTrafficState.queue1+currentTrafficState.queue2)-(prevTrafficState.queue1+prevTrafficState.queue2))/(prevTrafficState.queue1+prevTrafficState.queue2)
    #print("traffic state:",queue,flux)
    if queue>=0.7:
        currentTraffic=0
    elif queue>=0.3 and queue<0.7:
        currentTraffic=1
    elif queue<0.3:
        currentTraffic=2
    if flux<=0:
        trafficChange=0
    elif flux>0 and flux<0.5:
        trafficChange=1
    elif flux>0 and flux>=0.5:
        trafficChange=2
    #print(currentTraffic,trafficChange)
    return currentTraffic*3+trafficChange

def findMaxQvalueAction(state,filename):
    location = (filename)
    openWb = xlrd.open_workbook(location)
    sheet = openWb.sheet_by_index(0)
    max=-1
    for i in range(12):
        if sheet.cell_value(state,i)>max:
            max=sheet.cell_value(state,i)
            action=i+1

    return action

def calculateReward(currentTrafficState,prevTrafficState):
    #print("reward:")
    #print(currentTrafficState.queue1,currentTrafficState.queue2,prevTrafficState.queue1,prevTrafficState.queue2)
    queueEffect =(-(ceil(currentTrafficState.queue1)+ceil(currentTrafficState.queue2))+(ceil(prevTrafficState.queue1)+ceil(prevTrafficState.queue2)))*100
    #print("reward->",queueEffect)
    return queueEffect

def findRow(fileName):
    xfile = openpyxl.load_workbook(fileName)
    sheet = xfile.worksheets[0]
    for cell in sheet["A"]:
        if cell.value is None:
            return (cell.row)
            break
    else:
        return (cell.row + 1)

def calculatePhaseForA(action,queue1,queue2,oldTime1,oldTime2,avgDownstream):

    location2 = ("PreviousState2.xlsx")
    locationB = ("traffic2.xls")
    openWbB = xlrd.open_workbook(locationB)
    sheetB = openWbB.sheet_by_index(0)
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)
    factor1 = ceil(queue1) - (floor(0.266 * (oldTime1)))
    factor2 = ceil(queue2) - (floor(0.266 * (oldTime2)))
    sum=0
    if sheet2.nrows==0:
        for i in range(sheetB.nrows):
            sum = sum + int(sheetB.cell_value(i, 7))
        previousDownsream=sum
    else:
        previousDownsream=sheet2.cell_value(sheet2.nrows-1,6)
    #print(sheet2.nrows-1)
    newPhaseFor1 = oldTime1 + (action/12) * factor1 + (action/6)*(avgDownstream-previousDownsream)
    newPhaseFor2 = oldTime2 + (action/12) * factor2
    #print("calculatePhase")
    #print("factor1",factor1,"factor2",factor2)
    if newPhaseFor1<10:
        newPhaseFor1=10
    elif newPhaseFor1>120:
        newPhaseFor1=120
    if newPhaseFor2<10:
        newPhaseFor2=10
    elif newPhaseFor2>120:
        newPhaseFor2=120
    return newPhaseFor1,newPhaseFor2

def calculatePhaseForB(action,queue1,queue2,oldTime1,oldTime2,avgUpstream):
    location2A=("PreviousState.xlsx")
    openWb2A = xlrd.open_workbook(location2A)
    sheet2A = openWb2A.sheet_by_index(0)
    factor1 = ceil(queue1) - (floor(0.266 * (oldTime1)))
    factor2 = ceil(queue2) - (floor(0.266 * (oldTime2)))
    newPhaseFor1 = oldTime1 + (action/12) * factor1 + (action/6)*(avgUpstream-sheet2A.cell_value(sheet2A.nrows-1,6))
    newPhaseFor2 = oldTime2 + (action/12) * factor2
    #print("calculatePhase")
    #print("factor1",factor1,"factor2",factor2)
    if newPhaseFor1<10:
        newPhaseFor1=10
    elif newPhaseFor1>120:
        newPhaseFor1=120
    if newPhaseFor2<10:
        newPhaseFor2=10
    elif newPhaseFor2>120:
        newPhaseFor2=120
    return newPhaseFor1,newPhaseFor2

def updateA(state,currentTrafficState,prevTrafficState):
    file = openpyxl.load_workbook('QMatrix.xlsx')
    sheet = file['Sheet1']
    location2 = ("QMatrix.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)
    location3 = ("PreviousState.xlsx")
    openWb3 = xlrd.open_workbook(location3)
    sheet3 = openWb3.sheet_by_index(0)
    index = int((sheet3.cell_value(sheet3.nrows-1, 0) + 1))
    #print((int(sheet3.cell_value(sheet3.nrows-1, 5)+65)))
    c = chr(int(sheet3.cell_value(sheet3.nrows-1, 5)+64)) + str(index)
    #print("q",c)
    reward=calculateReward(currentTrafficState,prevTrafficState)
    #print(reward)
    m = []
    for i in range(12):
        m.append(sheet2.cell_value(state,i))
    #print((sheet3.cell_value(sheet3.nrows-1, 5)))
   # print(alpha*(reward+gamma*max(m))+(1-alpha)*int(sheet[c].value))
    #print(sheet[c].value)
    sheet[c]=alpha*(reward+gamma*max(m))+(1-alpha)*int(sheet[c].value)
    file.save('QMatrix.xlsx')

def updateB(state,currentTrafficState,prevTrafficState):
    file = openpyxl.load_workbook('QMatrix2.xlsx')
    sheet = file['Sheet1']
    location2 = ("QMatrix2.xlsx")
    openWb2 = xlrd.open_workbook(location2)
    sheet2 = openWb2.sheet_by_index(0)
    location3 = ("PreviousState2.xlsx")
    openWb3 = xlrd.open_workbook(location3)
    sheet3 = openWb3.sheet_by_index(0)
    index = int((sheet3.cell_value(sheet3.nrows-1, 0) + 1))
    #print(chr(sheet3.cell_value(sheet3.nrows-1, 5)+65))
    c = chr(int(sheet3.cell_value(sheet3.nrows-1, 5)+64)) + str(index)
    #print("q",c)
    reward=calculateReward(currentTrafficState,prevTrafficState)
    #print(reward)
    m = []
    for i in range(12):
        m.append(sheet2.cell_value(state,i))
    #print((sheet3.cell_value(sheet3.nrows-1, 5)))
   # print(alpha*(reward+gamma*max(m))+(1-alpha)*int(sheet[c].value))
    sheet[c]=alpha*(reward+gamma*max(m))+(1-alpha)*int(sheet[c].value)
    file.save('QMatrix2.xlsx')

